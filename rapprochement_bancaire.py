"""
=============================================================================
AUTOMATISATION COMPTABLE & RAPPROCHEMENT BANCAIRE
PME Agro-alimentaire
=============================================================================
Fonctionnalités :
  - Parsing de relevés bancaires PDF
  - Import Excel/CSV des données comptables
  - Rapprochement bancaire automatique (fuzzy matching)
  - Catégorisation intelligente des transactions
  - Détection d'anomalies
  - Génération de rapports Excel & PDF
  - Intégration Sage/EBP & API bancaire (Open Banking)
=============================================================================
Installation des dépendances :
  pip install pandas openpyxl pdfplumber fuzzywuzzy python-Levenshtein
              reportlab requests matplotlib seaborn python-dateutil tqdm
=============================================================================
"""

import os
import re
import json
import logging
import warnings
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
import numpy as np
import pdfplumber
import requests
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from fuzzywuzzy import fuzz, process
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, Image as RLImage
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from dateutil import parser as dateutil_parser

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# CONFIGURATION GLOBALE
# ─────────────────────────────────────────────

CONFIG = {
    "entreprise": "Agro PME Côte d'Ivoire",
    "siret": "123 456 789 00010",
    "devise": "FCFA",
    "symbole": "XOF",

    # Seuils de détection d'anomalies
    "seuil_montant_eleve": 5_000_000,       # FCFA
    "seuil_zscore_anomalie": 2.5,
    "seuil_doublon_jours": 3,
    "seuil_rapprochement_score": 80,         # Score fuzzy minimum

    # Chemins
    "dossier_entree": "data/entree",
    "dossier_sortie": "data/sortie",
    "dossier_logs": "logs",

    # API bancaire (Open Banking / Bridge API)
    "api_bancaire": {
        "actif": False,                      # Passer à True si API disponible
        "url_base": "https://api.bridgeapi.io/v2",
        "client_id": "VOTRE_CLIENT_ID",
        "client_secret": "VOTRE_CLIENT_SECRET",
    },

    # Sage / EBP export
    "logiciel_comptable": "Sage",            # ou "EBP"
    "format_export_sage": "CSV",
    "encodage_sage": "latin-1",
}

# ─────────────────────────────────────────────
# CATÉGORIES AGRO-ALIMENTAIRE
# ─────────────────────────────────────────────

CATEGORIES = {
    # Achats & Matières premières
    "Matières premières": [
        "cacao", "café", "anacarde", "noix", "igname", "manioc", "banane",
        "plantain", "tomate", "oignon", "riz", "maïs", "soja", "lait",
        "farine", "sucre", "huile", "sel", "épice", "condiment",
        "intrant", "semence", "engrais", "pesticide"
    ],
    "Emballages": [
        "sac", "carton", "bouteille", "bidon", "sachet", "emballage",
        "étiquette", "boîte", "conserve", "bocal", "plastique", "verre"
    ],
    "Transport & Logistique": [
        "transport", "livraison", "camion", "fret", "transit", "douane",
        "port", "colis", "chauffeur", "carburant", "essence", "gasoil"
    ],
    "Énergie & Utilities": [
        "électricité", "cie", "sodeci", "eau", "gaz", "groupe", "générateur",
        "fuel", "energie", "power", "utility"
    ],
    "Personnel": [
        "salaire", "paie", "cnps", "rts", "conge", "prime", "indemnite",
        "formation", "recrutement", "personnel", "employe", "ouvrier"
    ],
    "Équipements & Maintenance": [
        "machine", "equipement", "materiel", "reparation", "maintenance",
        "piece", "outil", "entretien", "installation", "calibrage"
    ],
    "Charges financières": [
        "interet", "agios", "commission", "frais bancaire", "virement",
        "credit", "pret", "remboursement", "leasing"
    ],
    "Impôts & Taxes": [
        "dgi", "impot", "taxe", "tva", "bic", "patente", "timbre",
        "droit", "amende", "penalite", "fis", "tresor"
    ],
    "Ventes & Recettes": [
        "vente", "facture client", "reglement client", "acompte client",
        "export", "livraison client", "commande client", "recette"
    ],
    "Loyers & Immobilier": [
        "loyer", "bail", "location", "immeuble", "entrepot", "bureau",
        "usine", "terrain", "foncier"
    ],
    "Télécommunications": [
        "mtn", "orange", "moov", "wave", "mobile money", "om", "telephone",
        "internet", "abonnement", "communication"
    ],
    "Assurances": [
        "assurance", "saar", "nsia", "allianz", "axa", "prime assurance",
        "sinistre", "couverture"
    ],
    "Divers": []  # Catch-all
}

# ─────────────────────────────────────────────
# SETUP LOGGING
# ─────────────────────────────────────────────

def setup_logging() -> logging.Logger:
    os.makedirs(CONFIG["dossier_logs"], exist_ok=True)
    log_file = os.path.join(
        CONFIG["dossier_logs"],
        f"rapprochement_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    )
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()


# ═══════════════════════════════════════════════════════════════════
# MODULE 1 : PARSING RELEVÉS BANCAIRES PDF
# ═══════════════════════════════════════════════════════════════════

class ParserReleveBancairePDF:
    """
    Parse les relevés bancaires PDF et extrait les transactions.
    Compatible avec les formats courants des banques ivoiriennes
    (SGBCI, BICICI, BNI, Ecobank, Coris Bank, etc.)
    """

    # Patterns regex pour détecter les transactions
    PATTERNS_TRANSACTION = [
        # Format: DD/MM/YYYY | libellé | débit | crédit | solde
        r"(\d{2}/\d{2}/\d{4})\s+(.+?)\s+([\d\s,\.]+)?\s+([\d\s,\.]+)?\s+([\d\s,\.]+)",
        # Format: DD-MM-YY libellé montant D/C
        r"(\d{2}[-/]\d{2}[-/]\d{2,4})\s+(.+?)\s+([\d\.,]+)\s*([DC])?",
        # Format avec n° de chèque
        r"(\d{2}/\d{2}/\d{4})\s+(\d+\s+)?(.+?)\s+([\d\s]+,\d{2})\s+([\d\s]+,\d{2})?",
    ]

    PATTERNS_SOLDE = [
        r"(?:solde|balance|sold)[^\d]*([\d\s]+[,\.]\d{2})",
        r"(?:nouveau solde|new balance)[^\d]*([\d\s]+[,\.]\d{2})",
    ]

    def parse_pdf(self, chemin_pdf: str) -> pd.DataFrame:
        """Extrait les transactions d'un relevé PDF."""
        logger.info(f"Parsing PDF : {chemin_pdf}")
        transactions = []

        try:
            with pdfplumber.open(chemin_pdf) as pdf:
                texte_complet = ""
                for page in pdf.pages:
                    # Essayer d'abord l'extraction de tableaux
                    tables = page.extract_tables()
                    if tables:
                        for table in tables:
                            df_table = self._traiter_tableau(table)
                            if df_table is not None:
                                transactions.append(df_table)

                    # Extraction texte en complément
                    texte = page.extract_text() or ""
                    texte_complet += texte + "\n"

                # Si pas de tableau structuré, parser le texte
                if not transactions:
                    df_texte = self._parser_texte(texte_complet)
                    if df_texte is not None and not df_texte.empty:
                        transactions.append(df_texte)

        except Exception as e:
            logger.error(f"Erreur parsing PDF {chemin_pdf}: {e}")
            return pd.DataFrame()

        if not transactions:
            logger.warning(f"Aucune transaction trouvée dans {chemin_pdf}")
            return pd.DataFrame()

        df = pd.concat(transactions, ignore_index=True)
        df = self._normaliser_dataframe(df)
        logger.info(f"  → {len(df)} transactions extraites")
        return df

    def _traiter_tableau(self, table: list) -> Optional[pd.DataFrame]:
        """Traite un tableau extrait par pdfplumber."""
        if not table or len(table) < 2:
            return None

        # Nettoyer le tableau
        rows = []
        for row in table:
            if row and any(cell for cell in row if cell):
                rows.append([str(cell or "").strip() for cell in row])

        if len(rows) < 2:
            return None

        # Identifier les colonnes
        headers = rows[0]
        col_map = self._identifier_colonnes(headers)

        if not col_map.get("date"):
            return None

        data = []
        for row in rows[1:]:
            try:
                item = self._extraire_ligne(row, col_map, headers)
                if item:
                    data.append(item)
            except Exception:
                continue

        return pd.DataFrame(data) if data else None

    def _identifier_colonnes(self, headers: list) -> dict:
        """Identifie automatiquement les colonnes date, libellé, débit, crédit."""
        col_map = {}
        keywords = {
            "date": ["date", "jour", "day"],
            "libelle": ["libelle", "libellé", "operation", "opération",
                        "description", "motif", "detail", "désignation"],
            "debit": ["debit", "débit", "retrait", "sortie", "montant debit", "-"],
            "credit": ["credit", "crédit", "versement", "entree", "entrée",
                       "montant credit", "+"],
            "solde": ["solde", "balance", "cumul"],
        }

        for idx, header in enumerate(headers):
            h = header.lower().strip()
            for col_type, mots in keywords.items():
                if any(mot in h for mot in mots):
                    if col_type not in col_map:
                        col_map[col_type] = idx

        return col_map

    def _extraire_ligne(self, row: list, col_map: dict, headers: list) -> Optional[dict]:
        """Extrait une transaction depuis une ligne de tableau."""
        if len(row) <= max(col_map.values(), default=0):
            return None

        date_str = row[col_map["date"]] if "date" in col_map else ""
        if not date_str or not re.search(r"\d{2}[/\-]\d{2}", date_str):
            return None

        date = self._parser_date(date_str)
        if not date:
            return None

        libelle = row[col_map["libelle"]] if "libelle" in col_map else ""
        debit = self._parser_montant(row[col_map["debit"]]) if "debit" in col_map else 0.0
        credit = self._parser_montant(row[col_map["credit"]]) if "credit" in col_map else 0.0
        solde = self._parser_montant(row[col_map["solde"]]) if "solde" in col_map else None

        # Montant net (débit négatif, crédit positif)
        montant = credit - debit if (credit or debit) else 0.0

        return {
            "date": date,
            "libelle": libelle.strip(),
            "debit": debit,
            "credit": credit,
            "montant": montant,
            "solde": solde,
            "source": "PDF",
        }

    def _parser_texte(self, texte: str) -> Optional[pd.DataFrame]:
        """Parse le texte brut du relevé en fallback."""
        lignes = texte.split("\n")
        data = []

        for ligne in lignes:
            for pattern in self.PATTERNS_TRANSACTION:
                match = re.search(pattern, ligne, re.IGNORECASE)
                if match:
                    groups = match.groups()
                    try:
                        date = self._parser_date(groups[0])
                        if not date:
                            continue
                        libelle = groups[1] if len(groups) > 1 else ""
                        montant_str = groups[2] if len(groups) > 2 else "0"
                        montant = self._parser_montant(montant_str or "0")

                        # Détecter D/C si présent
                        sens = groups[-1] if groups[-1] in ("D", "C") else None
                        if sens == "D":
                            montant = -abs(montant)
                        elif sens == "C":
                            montant = abs(montant)

                        data.append({
                            "date": date,
                            "libelle": libelle.strip(),
                            "debit": abs(montant) if montant < 0 else 0,
                            "credit": montant if montant > 0 else 0,
                            "montant": montant,
                            "solde": None,
                            "source": "PDF_texte",
                        })
                        break
                    except Exception:
                        continue

        return pd.DataFrame(data) if data else None

    @staticmethod
    def _parser_date(date_str: str) -> Optional[datetime]:
        """Parse une date dans divers formats."""
        if not date_str:
            return None
        try:
            date_str = re.sub(r"[^\d/\-\.]", " ", date_str).strip()
            return dateutil_parser.parse(date_str, dayfirst=True)
        except Exception:
            return None

    @staticmethod
    def _parser_montant(montant_str: str) -> float:
        """Convertit une chaîne montant en float."""
        if not montant_str or str(montant_str).strip() in ("", "-", "–", "—"):
            return 0.0
        try:
            s = str(montant_str)
            s = re.sub(r"[^\d,\.]", "", s)
            if not s:
                return 0.0
            # Gérer séparateurs décimaux
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                parts = s.split(",")
                if len(parts[-1]) <= 2:
                    s = s.replace(",", ".")
                else:
                    s = s.replace(",", "")
            return float(s)
        except Exception:
            return 0.0

    def _normaliser_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Normalise et nettoie le DataFrame des transactions."""
        if df.empty:
            return df

        colonnes_requises = ["date", "libelle", "debit", "credit", "montant", "solde", "source"]
        for col in colonnes_requises:
            if col not in df.columns:
                df[col] = None

        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df["debit"] = pd.to_numeric(df["debit"], errors="coerce").fillna(0)
        df["credit"] = pd.to_numeric(df["credit"], errors="coerce").fillna(0)
        df["montant"] = pd.to_numeric(df["montant"], errors="coerce")

        # Recalculer montant si manquant
        mask = df["montant"].isna()
        df.loc[mask, "montant"] = df.loc[mask, "credit"] - df.loc[mask, "debit"]

        df = df.dropna(subset=["date"])
        df = df[df["libelle"].str.strip().str.len() > 0]
        df = df.sort_values("date").reset_index(drop=True)
        df["id_transaction"] = [f"BQ_{i+1:06d}" for i in range(len(df))]

        return df


# ═══════════════════════════════════════════════════════════════════
# MODULE 2 : IMPORT DONNÉES COMPTABLES (Excel/CSV/Sage/EBP)
# ═══════════════════════════════════════════════════════════════════

class ImporteurDonneesComptables:
    """
    Importe les données depuis Excel, CSV, exports Sage ou EBP.
    """

    def importer_fichier(self, chemin: str) -> pd.DataFrame:
        """Détecte le format et importe le fichier."""
        ext = Path(chemin).suffix.lower()
        logger.info(f"Import données comptables : {chemin}")

        if ext in (".xlsx", ".xls"):
            return self._importer_excel(chemin)
        elif ext == ".csv":
            return self._importer_csv(chemin)
        elif ext == ".txt":
            # Format export Sage/EBP
            logiciel = CONFIG.get("logiciel_comptable", "Sage")
            if logiciel == "EBP":
                return self._importer_ebp(chemin)
            return self._importer_sage(chemin)
        else:
            logger.warning(f"Format non reconnu : {ext}")
            return pd.DataFrame()

    def _importer_excel(self, chemin: str) -> pd.DataFrame:
        """Import fichier Excel."""
        try:
            # Essayer plusieurs feuilles
            xl = pd.ExcelFile(chemin)
            for sheet in xl.sheet_names:
                df = pd.read_excel(chemin, sheet_name=sheet)
                df_norm = self._normaliser_comptabilite(df, source=f"Excel:{sheet}")
                if df_norm is not None and not df_norm.empty:
                    logger.info(f"  → Feuille '{sheet}' : {len(df_norm)} écritures")
                    return df_norm
        except Exception as e:
            logger.error(f"Erreur import Excel : {e}")
        return pd.DataFrame()

    def _importer_csv(self, chemin: str) -> pd.DataFrame:
        """Import fichier CSV avec détection auto du séparateur."""
        for sep in [";", ",", "\t", "|"]:
            for enc in ["utf-8", "latin-1", "cp1252"]:
                try:
                    df = pd.read_csv(chemin, sep=sep, encoding=enc,
                                     low_memory=False, on_bad_lines="skip")
                    if df.shape[1] >= 3:
                        df_norm = self._normaliser_comptabilite(df, source="CSV")
                        if df_norm is not None and not df_norm.empty:
                            return df_norm
                except Exception:
                    continue
        return pd.DataFrame()

    def _importer_sage(self, chemin: str) -> pd.DataFrame:
        """Import export Sage (format texte délimité)."""
        try:
            df = pd.read_csv(
                chemin,
                sep=";",
                encoding=CONFIG["encodage_sage"],
                on_bad_lines="skip",
                dtype=str,
            )
            # Mapping colonnes Sage standard
            mapping_sage = {
                "Date": "date", "DateEcheance": "date_echeance",
                "Libelle": "libelle", "Debit": "debit", "Credit": "credit",
                "CompteGeneral": "compte", "CompteAuxiliaire": "tiers",
                "NumeroDocument": "ref_document", "JournalCode": "journal",
                "Piece": "piece",
            }
            df = df.rename(columns={
                k: v for k, v in mapping_sage.items() if k in df.columns
            })
            return self._normaliser_comptabilite(df, source="Sage")
        except Exception as e:
            logger.error(f"Erreur import Sage : {e}")
            return pd.DataFrame()

    def _importer_ebp(self, chemin: str) -> pd.DataFrame:
        """Import export EBP Comptabilité."""
        try:
            df = pd.read_csv(
                chemin, sep="\t", encoding="utf-8",
                on_bad_lines="skip", dtype=str
            )
            mapping_ebp = {
                "Date pièce": "date", "Libellé": "libelle",
                "Débit": "debit", "Crédit": "credit",
                "N° compte": "compte", "Code journal": "journal",
                "N° pièce": "piece",
            }
            df = df.rename(columns={
                k: v for k, v in mapping_ebp.items() if k in df.columns
            })
            return self._normaliser_comptabilite(df, source="EBP")
        except Exception as e:
            logger.error(f"Erreur import EBP : {e}")
            return pd.DataFrame()

    def _normaliser_comptabilite(
        self, df: pd.DataFrame, source: str = "Inconnu"
    ) -> Optional[pd.DataFrame]:
        """Normalise les données comptables."""
        if df is None or df.empty:
            return None

        df.columns = df.columns.str.lower().str.strip()

        # Mapping flexible des colonnes
        aliases = {
            "date": ["date", "date_piece", "date_comptable", "date operation",
                     "date écriture", "jour"],
            "libelle": ["libelle", "libellé", "description", "designation",
                        "designation", "detail", "motif", "intitule"],
            "debit": ["debit", "débit", "montant_debit", "montant debit",
                      "sortie", "charge"],
            "credit": ["credit", "crédit", "montant_credit", "montant credit",
                       "entree", "produit"],
            "compte": ["compte", "numcompte", "compte_general", "n° compte",
                       "numéro compte"],
            "ref_document": ["ref_document", "ref", "piece", "numero",
                             "numero_piece", "n° pièce", "facture"],
        }

        for cible, sources_possibles in aliases.items():
            if cible not in df.columns:
                for src in sources_possibles:
                    if src in df.columns:
                        df[cible] = df[src]
                        break

        # Vérification colonnes minimales
        if "date" not in df.columns or "libelle" not in df.columns:
            return None

        df["date"] = pd.to_datetime(df["date"], errors="coerce", dayfirst=True)
        df = df.dropna(subset=["date"])

        for col in ["debit", "credit"]:
            if col in df.columns:
                df[col] = (
                    df[col].astype(str)
                    .str.replace(r"[^\d,\.]", "", regex=True)
                    .str.replace(",", ".")
                    .replace("", "0")
                    .pipe(pd.to_numeric, errors="coerce")
                    .fillna(0)
                )
            else:
                df[col] = 0.0

        df["montant"] = df["credit"] - df["debit"]
        df["source_donnee"] = source

        if "compte" not in df.columns:
            df["compte"] = ""
        if "ref_document" not in df.columns:
            df["ref_document"] = ""

        df = df.sort_values("date").reset_index(drop=True)
        df["id_ecriture"] = [f"CPT_{i+1:06d}" for i in range(len(df))]

        logger.info(f"  → {source} : {len(df)} écritures comptables normalisées")
        return df


# ═══════════════════════════════════════════════════════════════════
# MODULE 3 : IMPORT VIA API BANCAIRE (Open Banking)
# ═══════════════════════════════════════════════════════════════════

class ImporteurAPIBancaire:
    """
    Récupère les transactions via API bancaire (Bridge API / Open Banking).
    À adapter selon la banque utilisée.
    """

    def __init__(self):
        self.cfg = CONFIG["api_bancaire"]
        self.token = None

    def authentifier(self) -> bool:
        """Authentification OAuth2."""
        if not self.cfg["actif"]:
            logger.info("API bancaire désactivée dans la configuration.")
            return False
        try:
            resp = requests.post(
                f"{self.cfg['url_base']}/authenticate",
                json={
                    "client_id": self.cfg["client_id"],
                    "client_secret": self.cfg["client_secret"],
                },
                timeout=30,
            )
            resp.raise_for_status()
            self.token = resp.json().get("access_token")
            logger.info("API bancaire : authentification réussie")
            return bool(self.token)
        except Exception as e:
            logger.error(f"Erreur authentification API : {e}")
            return False

    def recuperer_transactions(
        self, compte_id: str,
        date_debut: datetime,
        date_fin: datetime
    ) -> pd.DataFrame:
        """Récupère les transactions d'un compte bancaire via API."""
        if not self.token:
            if not self.authentifier():
                return pd.DataFrame()

        transactions = []
        page = 1

        while True:
            try:
                resp = requests.get(
                    f"{self.cfg['url_base']}/accounts/{compte_id}/transactions",
                    headers={"Authorization": f"Bearer {self.token}"},
                    params={
                        "since": date_debut.strftime("%Y-%m-%d"),
                        "until": date_fin.strftime("%Y-%m-%d"),
                        "limit": 100,
                        "page": page,
                    },
                    timeout=30,
                )
                resp.raise_for_status()
                data = resp.json()
                items = data.get("resources", data.get("transactions", []))

                if not items:
                    break

                for item in items:
                    transactions.append({
                        "date": pd.to_datetime(
                            item.get("date") or item.get("transaction_date")
                        ),
                        "libelle": item.get("label") or item.get("description", ""),
                        "montant": float(item.get("amount", 0)),
                        "debit": abs(float(item.get("amount", 0))) if float(item.get("amount", 0)) < 0 else 0,
                        "credit": float(item.get("amount", 0)) if float(item.get("amount", 0)) > 0 else 0,
                        "ref_externe": str(item.get("id", "")),
                        "source": "API",
                    })

                page += 1
                if page > data.get("pagination", {}).get("total_pages", 1):
                    break

            except Exception as e:
                logger.error(f"Erreur API bancaire page {page}: {e}")
                break

        df = pd.DataFrame(transactions) if transactions else pd.DataFrame()
        logger.info(f"API bancaire : {len(df)} transactions récupérées")
        return df


# ═══════════════════════════════════════════════════════════════════
# MODULE 4 : CATÉGORISATION DES TRANSACTIONS
# ═══════════════════════════════════════════════════════════════════

class CategoriseurTransactions:
    """
    Catégorise automatiquement les transactions par analyse du libellé.
    Utilise des règles métier agro-alimentaire + correspondance floue.
    """

    def __init__(self):
        self.categories = CATEGORIES

    def categoriser_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Catégorise toutes les transactions d'un DataFrame."""
        if df.empty:
            return df

        df["categorie"] = df["libelle"].apply(self._categoriser_libelle)
        df["sous_categorie"] = df.apply(self._affiner_categorie, axis=1)

        nb_categorises = (df["categorie"] != "Divers").sum()
        pct = nb_categorises / len(df) * 100
        logger.info(f"Catégorisation : {nb_categorises}/{len(df)} transactions ({pct:.1f}%)")
        return df

    def _categoriser_libelle(self, libelle: str) -> str:
        """Détermine la catégorie d'un libellé."""
        if not libelle:
            return "Divers"

        libelle_norm = libelle.lower().strip()
        libelle_norm = re.sub(r"[^a-zàáâãäçèéêëìíîïòóôõöùúûü\s]", " ", libelle_norm)

        # Recherche par mots-clés
        for categorie, mots_cles in self.categories.items():
            if not mots_cles:
                continue
            for mot in mots_cles:
                if mot.lower() in libelle_norm:
                    return categorie

        # Fuzzy matching si pas trouvé
        tous_mots = []
        mapping = {}
        for cat, mots in self.categories.items():
            for mot in mots:
                tous_mots.append(mot)
                mapping[mot] = cat

        if tous_mots:
            mots_libelle = libelle_norm.split()
            for mot in mots_libelle:
                if len(mot) < 4:
                    continue
                match = process.extractOne(
                    mot, tous_mots,
                    scorer=fuzz.ratio,
                    score_cutoff=82
                )
                if match:
                    return mapping[match[0]]

        return "Divers"

    def _affiner_categorie(self, row: pd.Series) -> str:
        """Affine la catégorisation avec le montant et le contexte."""
        cat = row.get("categorie", "Divers")
        montant = row.get("montant", 0)

        if cat == "Divers":
            if montant > 0:
                return "Recette non identifiée"
            else:
                return "Dépense non identifiée"
        return cat


# ═══════════════════════════════════════════════════════════════════
# MODULE 5 : RAPPROCHEMENT BANCAIRE AUTOMATIQUE
# ═══════════════════════════════════════════════════════════════════

class RapprochementBancaire:
    """
    Effectue le rapprochement entre le relevé bancaire et la comptabilité.
    Utilise le fuzzy matching sur le montant, la date et le libellé.
    """

    SEUIL_SCORE = CONFIG["seuil_rapprochement_score"]
    FENETRE_JOURS = 5

    def rapprocher(
        self,
        df_banque: pd.DataFrame,
        df_comptable: pd.DataFrame
    ) -> dict:
        """
        Rapproche les transactions bancaires avec les écritures comptables.

        Returns:
            dict avec clés:
              - rapprochees: DataFrame des paires rapprochées
              - non_rapprochees_banque: transactions banque sans correspondance
              - non_rapprochees_compta: écritures sans correspondance
              - taux_rapprochement: pourcentage
        """
        logger.info("Début du rapprochement bancaire...")

        if df_banque.empty or df_comptable.empty:
            logger.warning("Un des DataFrames est vide — rapprochement impossible.")
            return self._resultat_vide(df_banque, df_comptable)

        df_banque = df_banque.copy()
        df_comptable = df_comptable.copy()
        df_banque["rapproche"] = False
        df_comptable["rapproche"] = False

        paires = []
        indices_compta_utilises = set()

        for idx_bq, ligne_bq in df_banque.iterrows():
            meilleur_score = 0
            meilleur_idx_cpt = None

            date_bq = ligne_bq["date"]
            montant_bq = ligne_bq["montant"]

            # Filtrer les écritures dans la fenêtre temporelle
            date_min = date_bq - timedelta(days=self.FENETRE_JOURS)
            date_max = date_bq + timedelta(days=self.FENETRE_JOURS)

            candidats = df_comptable[
                (df_comptable["date"] >= date_min) &
                (df_comptable["date"] <= date_max) &
                (~df_comptable.index.isin(indices_compta_utilises)) &
                (~df_comptable["rapproche"])
            ]

            for idx_cpt, ligne_cpt in candidats.iterrows():
                score = self._calculer_score(ligne_bq, ligne_cpt)
                if score > meilleur_score:
                    meilleur_score = score
                    meilleur_idx_cpt = idx_cpt

            if meilleur_score >= self.SEUIL_SCORE and meilleur_idx_cpt is not None:
                df_banque.at[idx_bq, "rapproche"] = True
                df_comptable.at[meilleur_idx_cpt, "rapproche"] = True
                indices_compta_utilises.add(meilleur_idx_cpt)

                paires.append({
                    "id_transaction_banque": ligne_bq.get("id_transaction", idx_bq),
                    "id_ecriture_compta": df_comptable.at[meilleur_idx_cpt, "id_ecriture"],
                    "date_banque": date_bq,
                    "date_compta": df_comptable.at[meilleur_idx_cpt, "date"],
                    "libelle_banque": ligne_bq["libelle"],
                    "libelle_compta": df_comptable.at[meilleur_idx_cpt, "libelle"],
                    "montant_banque": montant_bq,
                    "montant_compta": df_comptable.at[meilleur_idx_cpt, "montant"],
                    "ecart_montant": abs(
                        montant_bq - df_comptable.at[meilleur_idx_cpt, "montant"]
                    ),
                    "ecart_jours": abs(
                        (date_bq - df_comptable.at[meilleur_idx_cpt, "date"]).days
                    ),
                    "score_rapprochement": meilleur_score,
                    "statut": "Rapproché",
                })

        df_rapprochees = pd.DataFrame(paires)
        df_non_bq = df_banque[~df_banque["rapproche"]].copy()
        df_non_cpt = df_comptable[~df_comptable["rapproche"]].copy()

        taux = len(paires) / max(len(df_banque), 1) * 100

        logger.info(
            f"Rapprochement terminé : {len(paires)}/{len(df_banque)} "
            f"({taux:.1f}%) | Non rapprochées banque: {len(df_non_bq)} "
            f"| Non rapprochées compta: {len(df_non_cpt)}"
        )

        return {
            "rapprochees": df_rapprochees,
            "non_rapprochees_banque": df_non_bq,
            "non_rapprochees_compta": df_non_cpt,
            "taux_rapprochement": taux,
        }

    def _calculer_score(self, ligne_bq: pd.Series, ligne_cpt: pd.Series) -> float:
        """Calcule le score de correspondance entre deux lignes."""
        # Score montant (pondération 50%)
        mt_bq = abs(ligne_bq.get("montant", 0))
        mt_cpt = abs(ligne_cpt.get("montant", 0))

        if mt_bq == 0 and mt_cpt == 0:
            score_montant = 100
        elif max(mt_bq, mt_cpt) == 0:
            score_montant = 0
        else:
            ecart_pct = abs(mt_bq - mt_cpt) / max(mt_bq, mt_cpt) * 100
            score_montant = max(0, 100 - ecart_pct * 5)

        # Score libellé (pondération 35%)
        lib_bq = str(ligne_bq.get("libelle", "")).lower()
        lib_cpt = str(ligne_cpt.get("libelle", "")).lower()
        score_libelle = fuzz.token_set_ratio(lib_bq, lib_cpt)

        # Score date (pondération 15%)
        date_bq = ligne_bq.get("date")
        date_cpt = ligne_cpt.get("date")
        if date_bq and date_cpt:
            ecart_j = abs((date_bq - date_cpt).days)
            score_date = max(0, 100 - ecart_j * 15)
        else:
            score_date = 50

        return score_montant * 0.50 + score_libelle * 0.35 + score_date * 0.15

    @staticmethod
    def _resultat_vide(df_banque, df_comptable):
        return {
            "rapprochees": pd.DataFrame(),
            "non_rapprochees_banque": df_banque,
            "non_rapprochees_compta": df_comptable,
            "taux_rapprochement": 0.0,
        }


# ═══════════════════════════════════════════════════════════════════
# MODULE 6 : DÉTECTION D'ANOMALIES
# ═══════════════════════════════════════════════════════════════════

class DetecteurAnomalies:
    """
    Détecte les anomalies financières dans les transactions :
    - Doublons potentiels
    - Montants inhabituels (Z-score)
    - Transactions hors horaires (weekends/jours fériés)
    - Virements ronds suspects
    - Séquences inhabituelles
    """

    JOURS_FERIES_CI = [
        "01-01", "01-04", "05-01", "08-08", "11-15",
        "12-25", "04-14", "04-17",  # Variables selon année
    ]

    def analyser(self, df: pd.DataFrame) -> pd.DataFrame:
        """Analyse un DataFrame et retourne les anomalies détectées."""
        if df.empty:
            return pd.DataFrame()

        anomalies = []
        anomalies.extend(self._detecter_doublons(df))
        anomalies.extend(self._detecter_montants_inhabituels(df))
        anomalies.extend(self._detecter_weekends(df))
        anomalies.extend(self._detecter_montants_ronds(df))
        anomalies.extend(self._detecter_serie_rapide(df))

        if not anomalies:
            logger.info("Aucune anomalie détectée.")
            return pd.DataFrame()

        df_anomalies = pd.DataFrame(anomalies)
        df_anomalies = df_anomalies.sort_values(
            ["severite", "date"], ascending=[False, True]
        ).reset_index(drop=True)

        nb_haute = (df_anomalies["severite"] == "Haute").sum()
        nb_moy = (df_anomalies["severite"] == "Moyenne").sum()
        logger.warning(
            f"Anomalies détectées : {len(df_anomalies)} total "
            f"({nb_haute} haute, {nb_moy} moyenne)"
        )
        return df_anomalies

    def _detecter_doublons(self, df: pd.DataFrame) -> list:
        """Détecte les doublons potentiels (même montant, dates proches)."""
        anomalies = []
        df_sorted = df.sort_values("date")

        for i, row in df_sorted.iterrows():
            fenetre = df_sorted[
                (df_sorted["date"] >= row["date"] - timedelta(days=CONFIG["seuil_doublon_jours"])) &
                (df_sorted["date"] <= row["date"] + timedelta(days=CONFIG["seuil_doublon_jours"])) &
                (df_sorted.index != i)
            ]
            for j, row2 in fenetre.iterrows():
                if j <= i:
                    continue
                if abs(row["montant"] - row2["montant"]) < 1:
                    score_sim = fuzz.token_set_ratio(
                        str(row["libelle"]).lower(),
                        str(row2["libelle"]).lower()
                    )
                    if score_sim > 70:
                        anomalies.append({
                            "type": "Doublon potentiel",
                            "severite": "Haute",
                            "date": row["date"],
                            "libelle": row["libelle"],
                            "montant": row["montant"],
                            "detail": (
                                f"Transaction similaire le {row2['date'].strftime('%d/%m/%Y')} "
                                f"| Similarité : {score_sim}%"
                            ),
                            "id_ref": row.get("id_transaction", str(i)),
                        })
        return anomalies

    def _detecter_montants_inhabituels(self, df: pd.DataFrame) -> list:
        """Détecte les montants statistiquement inhabituels (Z-score)."""
        anomalies = []
        if len(df) < 5:
            return anomalies

        montants = df["montant"].abs()
        moyenne = montants.mean()
        std = montants.std()

        if std == 0:
            return anomalies

        for _, row in df.iterrows():
            zscore = abs(abs(row["montant"]) - moyenne) / std
            if zscore > CONFIG["seuil_zscore_anomalie"]:
                severite = "Haute" if abs(row["montant"]) > CONFIG["seuil_montant_eleve"] else "Moyenne"
                anomalies.append({
                    "type": "Montant inhabituel",
                    "severite": severite,
                    "date": row["date"],
                    "libelle": row["libelle"],
                    "montant": row["montant"],
                    "detail": (
                        f"Z-score : {zscore:.2f} | "
                        f"Moyenne : {moyenne:,.0f} {CONFIG['devise']} | "
                        f"Écart-type : {std:,.0f}"
                    ),
                    "id_ref": row.get("id_transaction", ""),
                })
        return anomalies

    def _detecter_weekends(self, df: pd.DataFrame) -> list:
        """Signale les transactions un weekend ou jour férié."""
        anomalies = []
        for _, row in df.iterrows():
            date = row["date"]
            if not isinstance(date, pd.Timestamp):
                continue
            est_weekend = date.weekday() >= 5
            est_ferie = date.strftime("%m-%d") in self.JOURS_FERIES_CI

            if est_weekend or est_ferie:
                raison = "Weekend" if est_weekend else "Jour férié"
                anomalies.append({
                    "type": f"Transaction {raison}",
                    "severite": "Faible",
                    "date": date,
                    "libelle": row["libelle"],
                    "montant": row["montant"],
                    "detail": f"Opération un {date.strftime('%A %d/%m/%Y')}",
                    "id_ref": row.get("id_transaction", ""),
                })
        return anomalies

    def _detecter_montants_ronds(self, df: pd.DataFrame) -> list:
        """Détecte les montants ronds élevés (potentiellement suspects)."""
        anomalies = []
        seuil = CONFIG["seuil_montant_eleve"] * 0.2

        for _, row in df.iterrows():
            mt = abs(row["montant"])
            if mt >= seuil and mt % 100_000 == 0:
                anomalies.append({
                    "type": "Montant rond élevé",
                    "severite": "Faible",
                    "date": row["date"],
                    "libelle": row["libelle"],
                    "montant": row["montant"],
                    "detail": f"Montant rond de {mt:,.0f} {CONFIG['devise']}",
                    "id_ref": row.get("id_transaction", ""),
                })
        return anomalies

    def _detecter_serie_rapide(self, df: pd.DataFrame) -> list:
        """Détecte les séries de transactions rapides (même jour, même libellé)."""
        anomalies = []
        df_temp = df.copy()
        df_temp["date_j"] = pd.to_datetime(df_temp["date"]).dt.date
        df_temp["libelle_norm"] = df_temp["libelle"].str.lower().str.strip()

        groupes = df_temp.groupby(["date_j", "libelle_norm"])
        for (date_j, lib), groupe in groupes:
            if len(groupe) >= 3:
                anomalies.append({
                    "type": "Série transactions rapides",
                    "severite": "Moyenne",
                    "date": pd.Timestamp(date_j),
                    "libelle": lib,
                    "montant": groupe["montant"].sum(),
                    "detail": f"{len(groupe)} transactions identiques le même jour",
                    "id_ref": ";".join(groupe.get("id_transaction", groupe.index.astype(str)).tolist()),
                })
        return anomalies


# ═══════════════════════════════════════════════════════════════════
# MODULE 7 : GÉNÉRATION RAPPORT EXCEL
# ═══════════════════════════════════════════════════════════════════

class GenerateurRapportExcel:
    """
    Génère un classeur Excel professionnel avec :
    - Tableau de bord récapitulatif
    - Transactions bancaires
    - Écritures comptables
    - Rapprochement détaillé
    - Anomalies signalées
    - Graphiques analytiques
    """

    # Couleurs PME professionnelles
    COULEUR_ENTETE = "1B4F72"        # Bleu marine
    COULEUR_TITRE = "2E86C1"         # Bleu moyen
    COULEUR_SUCCES = "1E8449"        # Vert
    COULEUR_ALERTE = "CB4335"        # Rouge
    COULEUR_AVERTISSEMENT = "D68910" # Orange
    COULEUR_LIGNE_PAIRE = "EBF5FB"   # Bleu très clair
    COULEUR_LIGNE_IMPAIRE = "FDFEFE"  # Blanc cassé

    def generer(
        self,
        resultats_rapprochement: dict,
        df_banque: pd.DataFrame,
        df_comptable: pd.DataFrame,
        df_anomalies: pd.DataFrame,
        periode: str,
        chemin_sortie: str
    ) -> str:
        """Génère le rapport Excel complet."""
        logger.info(f"Génération rapport Excel : {chemin_sortie}")
        wb = Workbook()

        # Supprimer feuille par défaut
        wb.remove(wb.active)

        # Créer les feuilles
        self._creer_dashboard(wb, resultats_rapprochement, df_banque, df_anomalies, periode)
        self._creer_feuille_transactions(wb, df_banque, "Relevé Bancaire", "BQ")
        self._creer_feuille_transactions(wb, df_comptable, "Comptabilité", "CPT")
        self._creer_feuille_rapprochement(wb, resultats_rapprochement)
        if not df_anomalies.empty:
            self._creer_feuille_anomalies(wb, df_anomalies)
        self._creer_feuille_categories(wb, df_banque)

        os.makedirs(os.path.dirname(chemin_sortie) or ".", exist_ok=True)
        wb.save(chemin_sortie)
        logger.info(f"  → Rapport Excel enregistré : {chemin_sortie}")
        return chemin_sortie

    def _style_entete(self, gras=True, couleur_fond=None, taille=11,
                      couleur_texte="FFFFFF", alignement="center"):
        font = Font(bold=gras, color=couleur_texte, size=taille, name="Arial")
        fill = PatternFill("solid", fgColor=couleur_fond or self.COULEUR_ENTETE)
        align = Alignment(horizontal=alignement, vertical="center", wrap_text=True)
        return font, fill, align

    def _bordure_fine(self):
        thin = Side(style="thin", color="CCCCCC")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _creer_dashboard(self, wb, resultats, df_banque, df_anomalies, periode):
        ws = wb.create_sheet("📊 Tableau de Bord")
        ws.sheet_view.showGridLines = False

        # Titre principal
        ws.merge_cells("A1:H1")
        ws["A1"] = f"RAPPROCHEMENT BANCAIRE — {CONFIG['entreprise'].upper()}"
        ws["A1"].font = Font(bold=True, size=16, color="FFFFFF", name="Arial")
        ws["A1"].fill = PatternFill("solid", fgColor=self.COULEUR_ENTETE)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        ws.merge_cells("A2:H2")
        ws["A2"] = f"Période : {periode}  |  Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["A2"].font = Font(italic=True, size=10, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 20

        # KPIs
        taux = resultats.get("taux_rapprochement", 0)
        nb_rappr = len(resultats.get("rapprochees", pd.DataFrame()))
        nb_non_bq = len(resultats.get("non_rapprochees_banque", pd.DataFrame()))
        nb_non_cpt = len(resultats.get("non_rapprochees_compta", pd.DataFrame()))
        nb_anomalies = len(df_anomalies)
        nb_anomalies_hautes = (
            (df_anomalies["severite"] == "Haute").sum()
            if not df_anomalies.empty else 0
        )

        kpis = [
            ("Taux Rapprochement", f"{taux:.1f}%",
             self.COULEUR_SUCCES if taux >= 85 else self.COULEUR_AVERTISSEMENT),
            ("Transactions Rapprochées", str(nb_rappr), self.COULEUR_TITRE),
            ("Non Rappr. Banque", str(nb_non_bq),
             self.COULEUR_ALERTE if nb_non_bq > 0 else self.COULEUR_SUCCES),
            ("Non Rappr. Comptabilité", str(nb_non_cpt),
             self.COULEUR_ALERTE if nb_non_cpt > 0 else self.COULEUR_SUCCES),
            ("Anomalies Hautes", str(nb_anomalies_hautes),
             self.COULEUR_ALERTE if nb_anomalies_hautes > 0 else self.COULEUR_SUCCES),
            ("Total Anomalies", str(nb_anomalies), self.COULEUR_AVERTISSEMENT),
        ]

        ws.row_dimensions[4].height = 18
        ws.row_dimensions[5].height = 50
        ws.row_dimensions[6].height = 30
        ws.row_dimensions[7].height = 18

        for i, (label, valeur, couleur) in enumerate(kpis):
            col = i + 1
            col_lettre = get_column_letter(col)

            ws.merge_cells(f"{col_lettre}4:{col_lettre}4")
            ws[f"{col_lettre}4"] = label
            ws[f"{col_lettre}4"].font = Font(bold=True, size=9, color="555555")
            ws[f"{col_lettre}4"].alignment = Alignment(horizontal="center")

            ws[f"{col_lettre}5"] = valeur
            ws[f"{col_lettre}5"].font = Font(bold=True, size=22, color=couleur)
            ws[f"{col_lettre}5"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"{col_lettre}5"].fill = PatternFill("solid", fgColor="F8F9FA")
            ws[f"{col_lettre}5"].border = self._bordure_fine()

            ws.column_dimensions[col_lettre].width = 22

        # Résumé financier
        row = 9
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = "RÉSUMÉ FINANCIER"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=self.COULEUR_TITRE)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 22

        if not df_banque.empty:
            total_debits = df_banque["debit"].sum()
            total_credits = df_banque["credit"].sum()
            solde_net = total_credits - total_debits

            resume_data = [
                ("Total Débits (sorties)", total_debits, self.COULEUR_ALERTE),
                ("Total Crédits (entrées)", total_credits, self.COULEUR_SUCCES),
                ("Solde Net", solde_net,
                 self.COULEUR_SUCCES if solde_net >= 0 else self.COULEUR_ALERTE),
            ]

            for j, (label, valeur, couleur) in enumerate(resume_data):
                r = row + 1 + j
                ws[f"A{r}"] = label
                ws[f"A{r}"].font = Font(bold=True, size=10)
                ws[f"B{r}"] = valeur
                ws[f"B{r}"].font = Font(bold=True, size=11, color=couleur)
                ws[f"B{r}"].number_format = '#,##0 "XOF"'
                ws[f"C{r}"] = CONFIG["devise"]
                ws.row_dimensions[r].height = 18

        ws.sheet_properties.tabColor = "1B4F72"

    def _creer_feuille_transactions(
        self, wb, df: pd.DataFrame, titre: str, prefix: str
    ):
        """Crée une feuille de transactions (banque ou comptabilité)."""
        ws = wb.create_sheet(f"{'🏦' if prefix=='BQ' else '📒'} {titre}")
        ws.sheet_view.showGridLines = False

        if df.empty:
            ws["A1"] = "Aucune donnée disponible"
            return

        # Colonnes à afficher
        colonnes = {
            "date": "Date",
            "libelle": "Libellé",
            "debit": f"Débit ({CONFIG['devise']})",
            "credit": f"Crédit ({CONFIG['devise']})",
            "montant": f"Montant ({CONFIG['devise']})",
            "categorie": "Catégorie",
            "solde": f"Solde ({CONFIG['devise']})",
        }
        cols_dispo = {k: v for k, v in colonnes.items() if k in df.columns}

        # En-têtes
        font_h, fill_h, align_h = self._style_entete()
        for col_idx, (col_key, col_label) in enumerate(cols_dispo.items(), 1):
            cell = ws.cell(row=1, column=col_idx, value=col_label)
            cell.font = font_h
            cell.fill = fill_h
            cell.alignment = align_h
            cell.border = self._bordure_fine()

        ws.row_dimensions[1].height = 22

        # Données
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            couleur_fond = self.COULEUR_LIGNE_PAIRE if row_idx % 2 == 0 else self.COULEUR_LIGNE_IMPAIRE
            fill_ligne = PatternFill("solid", fgColor=couleur_fond)

            for col_idx, (col_key, _) in enumerate(cols_dispo.items(), 1):
                valeur = row.get(col_key, "")
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill_ligne
                cell.border = self._bordure_fine()

                if col_key == "date" and pd.notna(valeur):
                    cell.value = pd.Timestamp(valeur).to_pydatetime()
                    cell.number_format = "DD/MM/YYYY"
                    cell.alignment = Alignment(horizontal="center")
                elif col_key in ("debit", "credit", "montant", "solde"):
                    cell.value = float(valeur) if pd.notna(valeur) else 0
                    cell.number_format = '#,##0;(#,##0);"-"'
                    cell.alignment = Alignment(horizontal="right")
                    if col_key == "montant":
                        couleur_mt = (
                            self.COULEUR_SUCCES if (float(valeur) if pd.notna(valeur) else 0) > 0
                            else self.COULEUR_ALERTE
                        )
                        cell.font = Font(color=couleur_mt, size=10)
                elif col_key == "libelle":
                    cell.value = str(valeur)
                    cell.alignment = Alignment(wrap_text=False)
                else:
                    cell.value = str(valeur) if pd.notna(valeur) else ""

        # Largeurs de colonnes
        largeurs = {
            "date": 14, "libelle": 40, "debit": 16, "credit": 16,
            "montant": 16, "categorie": 22, "solde": 16
        }
        for col_idx, col_key in enumerate(cols_dispo.keys(), 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = largeurs.get(col_key, 15)

        # Figer la première ligne
        ws.freeze_panes = "A2"

        # Filtre auto
        derniere_col = get_column_letter(len(cols_dispo))
        derniere_ligne = len(df) + 1
        ws.auto_filter.ref = f"A1:{derniere_col}{derniere_ligne}"

    def _creer_feuille_rapprochement(self, wb, resultats: dict):
        """Crée la feuille de rapprochement avec code couleur."""
        ws = wb.create_sheet("🔗 Rapprochement")
        ws.sheet_view.showGridLines = False

        df_rappr = resultats.get("rapprochees", pd.DataFrame())
        df_non_bq = resultats.get("non_rapprochees_banque", pd.DataFrame())

        row = 1

        # Section : Transactions rapprochées
        ws.merge_cells(f"A{row}:I{row}")
        ws[f"A{row}"] = f"✅ TRANSACTIONS RAPPROCHÉES ({len(df_rappr)})"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=self.COULEUR_SUCCES)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 22
        row += 1

        if not df_rappr.empty:
            cols_rappr = [
                ("date_banque", "Date Banque", 14),
                ("libelle_banque", "Libellé Banque", 35),
                ("montant_banque", "Montant Banque", 18),
                ("date_compta", "Date Compta", 14),
                ("libelle_compta", "Libellé Compta", 35),
                ("montant_compta", "Montant Compta", 18),
                ("ecart_montant", "Écart Montant", 16),
                ("ecart_jours", "Écart Jours", 12),
                ("score_rapprochement", "Score", 10),
            ]

            font_h, fill_h, align_h = self._style_entete(couleur_fond=self.COULEUR_SUCCES)
            for col_idx, (col_key, label, largeur) in enumerate(cols_rappr, 1):
                cell = ws.cell(row=row, column=col_idx, value=label)
                cell.font = font_h
                cell.fill = fill_h
                cell.alignment = align_h
                cell.border = self._bordure_fine()
                ws.column_dimensions[get_column_letter(col_idx)].width = largeur
            row += 1

            for r_idx, (_, r) in enumerate(df_rappr.iterrows()):
                couleur_fond = "E9F7EF" if r_idx % 2 == 0 else "FDFFFE"
                fill = PatternFill("solid", fgColor=couleur_fond)
                for col_idx, (col_key, _, _) in enumerate(cols_rappr, 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.fill = fill
                    cell.border = self._bordure_fine()
                    val = r.get(col_key, "")
                    if "date" in col_key and pd.notna(val):
                        cell.value = pd.Timestamp(val).to_pydatetime()
                        cell.number_format = "DD/MM/YYYY"
                        cell.alignment = Alignment(horizontal="center")
                    elif "montant" in col_key or col_key == "ecart_montant":
                        cell.value = float(val) if pd.notna(val) else 0
                        cell.number_format = '#,##0;(#,##0)'
                        cell.alignment = Alignment(horizontal="right")
                    elif col_key == "score_rapprochement":
                        cell.value = float(val) if pd.notna(val) else 0
                        cell.number_format = "0.0"
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.value = str(val) if pd.notna(val) else ""
                row += 1

        # Section : Non rapprochées
        row += 1
        ws.merge_cells(f"A{row}:I{row}")
        ws[f"A{row}"] = f"❌ NON RAPPROCHÉES BANQUE ({len(df_non_bq)})"
        ws[f"A{row}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor=self.COULEUR_ALERTE)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].height = 22
        row += 1

        if not df_non_bq.empty:
            cols_non = ["date", "libelle", "montant", "debit", "credit", "categorie"]
            cols_dispo = [c for c in cols_non if c in df_non_bq.columns]
            font_h, fill_h, align_h = self._style_entete(couleur_fond=self.COULEUR_ALERTE)
            for col_idx, col_key in enumerate(cols_dispo, 1):
                cell = ws.cell(row=row, column=col_idx, value=col_key.replace("_", " ").title())
                cell.font = font_h
                cell.fill = fill_h
                cell.alignment = align_h
                cell.border = self._bordure_fine()
            row += 1

            for _, r in df_non_bq.iterrows():
                for col_idx, col_key in enumerate(cols_dispo, 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.fill = PatternFill("solid", fgColor="FDEDEC")
                    cell.border = self._bordure_fine()
                    val = r.get(col_key, "")
                    if col_key == "date" and pd.notna(val):
                        cell.value = pd.Timestamp(val).to_pydatetime()
                        cell.number_format = "DD/MM/YYYY"
                    elif col_key in ("montant", "debit", "credit"):
                        cell.value = float(val) if pd.notna(val) else 0
                        cell.number_format = '#,##0;(#,##0)'
                    else:
                        cell.value = str(val) if pd.notna(val) else ""
                row += 1

        ws.freeze_panes = "A2"

    def _creer_feuille_anomalies(self, wb, df_anomalies: pd.DataFrame):
        """Crée la feuille d'anomalies avec code de sévérité."""
        ws = wb.create_sheet("⚠️ Anomalies")
        ws.sheet_view.showGridLines = False

        couleurs_severite = {
            "Haute": ("FDEDEC", "CB4335"),
            "Moyenne": ("FEF9E7", "D68910"),
            "Faible": ("EAFAF1", "1E8449"),
        }

        entetes = ["Sévérité", "Type", "Date", "Libellé", "Montant", "Détail"]
        font_h, fill_h, align_h = self._style_entete()
        for col_idx, entete in enumerate(entetes, 1):
            cell = ws.cell(row=1, column=col_idx, value=entete)
            cell.font = font_h
            cell.fill = fill_h
            cell.alignment = align_h
            cell.border = self._bordure_fine()

        largeurs = [12, 30, 14, 35, 18, 50]
        for i, l in enumerate(largeurs, 1):
            ws.column_dimensions[get_column_letter(i)].width = l

        for row_idx, (_, row) in enumerate(df_anomalies.iterrows(), 2):
            sev = row.get("severite", "Faible")
            fond, texte_couleur = couleurs_severite.get(sev, ("FFFFFF", "000000"))
            fill = PatternFill("solid", fgColor=fond)

            valeurs = [
                sev,
                row.get("type", ""),
                row.get("date"),
                row.get("libelle", ""),
                row.get("montant", 0),
                row.get("detail", ""),
            ]

            for col_idx, val in enumerate(valeurs, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.border = self._bordure_fine()

                if col_idx == 1:  # Sévérité
                    cell.value = val
                    cell.font = Font(bold=True, color=texte_couleur)
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx == 3:  # Date
                    if pd.notna(val):
                        cell.value = pd.Timestamp(val).to_pydatetime()
                        cell.number_format = "DD/MM/YYYY"
                        cell.alignment = Alignment(horizontal="center")
                elif col_idx == 5:  # Montant
                    cell.value = float(val) if pd.notna(val) else 0
                    cell.number_format = '#,##0;(#,##0)'
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = str(val) if pd.notna(val) else ""
                    if col_idx == 6:
                        cell.alignment = Alignment(wrap_text=True)

            ws.row_dimensions[row_idx].height = 18

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:F{len(df_anomalies)+1}"

    def _creer_feuille_categories(self, wb, df_banque: pd.DataFrame):
        """Crée la feuille d'analyse par catégories avec graphique."""
        ws = wb.create_sheet("📈 Catégories")
        ws.sheet_view.showGridLines = False

        if df_banque.empty or "categorie" not in df_banque.columns:
            ws["A1"] = "Données insuffisantes"
            return

        analyse = (
            df_banque.groupby("categorie")
            .agg(
                nombre=("montant", "count"),
                total_debit=("debit", "sum"),
                total_credit=("credit", "sum"),
                montant_net=("montant", "sum"),
            )
            .sort_values("total_debit", ascending=False)
            .reset_index()
        )

        entetes = ["Catégorie", "Nb Transactions",
                   f"Débits ({CONFIG['devise']})",
                   f"Crédits ({CONFIG['devise']})",
                   f"Net ({CONFIG['devise']})"]

        font_h, fill_h, align_h = self._style_entete(couleur_fond=self.COULEUR_TITRE)
        for col_idx, entete in enumerate(entetes, 1):
            cell = ws.cell(row=1, column=col_idx, value=entete)
            cell.font = font_h
            cell.fill = fill_h
            cell.alignment = align_h
            cell.border = self._bordure_fine()

        for row_idx, (_, row) in enumerate(analyse.iterrows(), 2):
            couleur_fond = self.COULEUR_LIGNE_PAIRE if row_idx % 2 == 0 else self.COULEUR_LIGNE_IMPAIRE
            fill = PatternFill("solid", fgColor=couleur_fond)
            valeurs = [
                row["categorie"], row["nombre"],
                row["total_debit"], row["total_credit"], row["montant_net"]
            ]
            for col_idx, val in enumerate(valeurs, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
                cell.border = self._bordure_fine()
                if col_idx >= 3:
                    cell.number_format = '#,##0;(#,##0)'
                    cell.alignment = Alignment(horizontal="right")

        largeurs = [28, 18, 20, 20, 20]
        for i, l in enumerate(largeurs, 1):
            ws.column_dimensions[get_column_letter(i)].width = l

        # Graphique barres
        if len(analyse) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Dépenses par Catégorie"
            chart.y_axis.title = f"Montant ({CONFIG['devise']})"
            chart.x_axis.title = "Catégorie"
            chart.style = 10
            chart.width = 20
            chart.height = 12

            data = Reference(ws, min_col=3, min_row=1,
                             max_row=len(analyse) + 1, max_col=3)
            cats = Reference(ws, min_col=1, min_row=2,
                             max_row=len(analyse) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "G2")

        ws.freeze_panes = "A2"


# ═══════════════════════════════════════════════════════════════════
# MODULE 8 : GÉNÉRATION RAPPORT PDF
# ═══════════════════════════════════════════════════════════════════

class GenerateurRapportPDF:
    """
    Génère un rapport PDF exécutif du rapprochement bancaire.
    """

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._definir_styles()

    def _definir_styles(self):
        self.style_titre = ParagraphStyle(
            "Titre", parent=self.styles["Title"],
            fontSize=18, textColor=colors.HexColor("#1B4F72"),
            spaceAfter=6, alignment=TA_CENTER
        )
        self.style_sous_titre = ParagraphStyle(
            "SousTitre", parent=self.styles["Normal"],
            fontSize=11, textColor=colors.HexColor("#555555"),
            spaceAfter=12, alignment=TA_CENTER
        )
        self.style_section = ParagraphStyle(
            "Section", parent=self.styles["Heading2"],
            fontSize=13, textColor=colors.white,
            backColor=colors.HexColor("#2E86C1"),
            spaceAfter=8, spaceBefore=14,
            leftIndent=6,
        )
        self.style_normal = ParagraphStyle(
            "Normal2", parent=self.styles["Normal"],
            fontSize=10, spaceAfter=4
        )
        self.style_alerte = ParagraphStyle(
            "Alerte", parent=self.styles["Normal"],
            fontSize=10, textColor=colors.HexColor("#CB4335"),
            spaceAfter=4
        )

    def generer(
        self,
        resultats_rapprochement: dict,
        df_banque: pd.DataFrame,
        df_anomalies: pd.DataFrame,
        periode: str,
        chemin_sortie: str
    ) -> str:
        """Génère le rapport PDF."""
        logger.info(f"Génération rapport PDF : {chemin_sortie}")
        os.makedirs(os.path.dirname(chemin_sortie) or ".", exist_ok=True)

        doc = SimpleDocTemplate(
            chemin_sortie,
            pagesize=A4,
            rightMargin=2*cm, leftMargin=2*cm,
            topMargin=2*cm, bottomMargin=2*cm,
        )

        contenu = []
        contenu.extend(self._page_garde(periode))
        contenu.append(HRFlowable(width="100%", thickness=2,
                                   color=colors.HexColor("#1B4F72")))
        contenu.append(Spacer(1, 0.4*cm))
        contenu.extend(self._section_resume(resultats_rapprochement, df_banque))
        contenu.extend(self._section_rapprochement(resultats_rapprochement))
        contenu.extend(self._section_anomalies(df_anomalies))
        if not df_banque.empty and "categorie" in df_banque.columns:
            contenu.extend(self._section_categories(df_banque))
        contenu.extend(self._pied_page())

        doc.build(contenu)
        logger.info(f"  → Rapport PDF enregistré : {chemin_sortie}")
        return chemin_sortie

    def _page_garde(self, periode: str) -> list:
        elements = [
            Spacer(1, 1*cm),
            Paragraph(CONFIG["entreprise"].upper(), self.style_titre),
            Paragraph("Rapport de Rapprochement Bancaire", self.style_sous_titre),
            Paragraph(f"Période : <b>{periode}</b>", self.style_sous_titre),
            Paragraph(
                f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}",
                ParagraphStyle("Date", parent=self.style_normal,
                               alignment=TA_CENTER, textColor=colors.grey)
            ),
            Spacer(1, 0.8*cm),
        ]
        return elements

    def _section_resume(self, resultats: dict, df_banque: pd.DataFrame) -> list:
        elements = [
            Paragraph("1. RÉSUMÉ EXÉCUTIF", self.style_section),
            Spacer(1, 0.3*cm),
        ]

        taux = resultats.get("taux_rapprochement", 0)
        nb_rappr = len(resultats.get("rapprochees", pd.DataFrame()))
        nb_non_bq = len(resultats.get("non_rapprochees_banque", pd.DataFrame()))
        nb_non_cpt = len(resultats.get("non_rapprochees_compta", pd.DataFrame()))

        kpi_data = [
            ["Indicateur", "Valeur", "Statut"],
            ["Taux de rapprochement", f"{taux:.1f}%",
             "✅ Bon" if taux >= 85 else "⚠️ À améliorer"],
            ["Transactions rapprochées", str(nb_rappr), "—"],
            ["Non rapprochées (banque)", str(nb_non_bq),
             "❌ Action requise" if nb_non_bq > 0 else "✅"],
            ["Non rapprochées (compta)", str(nb_non_cpt),
             "❌ Action requise" if nb_non_cpt > 0 else "✅"],
        ]

        if not df_banque.empty:
            kpi_data += [
                ["Total Débits", f"{df_banque['debit'].sum():,.0f} {CONFIG['devise']}", "—"],
                ["Total Crédits", f"{df_banque['credit'].sum():,.0f} {CONFIG['devise']}", "—"],
            ]

        table = Table(kpi_data, colWidths=[7*cm, 5*cm, 5*cm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B4F72")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#EBF5FB"), colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))
        return elements

    def _section_rapprochement(self, resultats: dict) -> list:
        elements = [
            Paragraph("2. DÉTAIL DU RAPPROCHEMENT", self.style_section),
            Spacer(1, 0.3*cm),
        ]

        df_non_bq = resultats.get("non_rapprochees_banque", pd.DataFrame())

        if not df_non_bq.empty:
            elements.append(Paragraph(
                f"<b>Transactions bancaires sans correspondance comptable "
                f"({len(df_non_bq)}) :</b>",
                self.style_alerte
            ))

            cols = ["date", "libelle", "montant"]
            cols_dispo = [c for c in cols if c in df_non_bq.columns]
            data = [["Date", "Libellé", f"Montant ({CONFIG['devise']})"]]

            for _, row in df_non_bq.head(20).iterrows():
                ligne = []
                for c in cols_dispo:
                    val = row.get(c, "")
                    if c == "date" and pd.notna(val):
                        ligne.append(pd.Timestamp(val).strftime("%d/%m/%Y"))
                    elif c == "montant":
                        ligne.append(f"{float(val):,.0f}" if pd.notna(val) else "0")
                    else:
                        txt = str(val)[:50] if pd.notna(val) else ""
                        ligne.append(txt)
                data.append(ligne)

            if len(df_non_bq) > 20:
                data.append([f"... et {len(df_non_bq)-20} autres", "", ""])

            table = Table(data, colWidths=[3*cm, 9*cm, 4.5*cm])
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#CB4335")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.HexColor("#FDEDEC"), colors.white]),
                ("ALIGN", (2, 0), (2, -1), "RIGHT"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph(
                "✅ Toutes les transactions bancaires ont été rapprochées.",
                self.style_normal
            ))

        elements.append(Spacer(1, 0.5*cm))
        return elements

    def _section_anomalies(self, df_anomalies: pd.DataFrame) -> list:
        elements = [
            Paragraph("3. ANOMALIES DÉTECTÉES", self.style_section),
            Spacer(1, 0.3*cm),
        ]

        if df_anomalies.empty:
            elements.append(Paragraph("✅ Aucune anomalie détectée.", self.style_normal))
            elements.append(Spacer(1, 0.5*cm))
            return elements

        nb_haute = (df_anomalies["severite"] == "Haute").sum()
        nb_moy = (df_anomalies["severite"] == "Moyenne").sum()
        nb_faible = (df_anomalies["severite"] == "Faible").sum()

        elements.append(Paragraph(
            f"<b>{len(df_anomalies)} anomalie(s) détectée(s)</b> : "
            f"{nb_haute} haute(s), {nb_moy} moyenne(s), {nb_faible} faible(s)",
            self.style_alerte if nb_haute > 0 else self.style_normal
        ))
        elements.append(Spacer(1, 0.2*cm))

        data = [["Sévérité", "Type", "Date", "Montant", "Détail"]]
        for _, row in df_anomalies.head(15).iterrows():
            date_str = ""
            if pd.notna(row.get("date")):
                date_str = pd.Timestamp(row["date"]).strftime("%d/%m/%Y")
            montant_str = ""
            if pd.notna(row.get("montant")):
                montant_str = f"{float(row['montant']):,.0f}"
            data.append([
                str(row.get("severite", "")),
                str(row.get("type", ""))[:30],
                date_str,
                montant_str,
                str(row.get("detail", ""))[:60],
            ])

        table = Table(data, colWidths=[2.2*cm, 4*cm, 2.5*cm, 3*cm, 5*cm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D68910")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#FEF9E7"), colors.white]),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))
        return elements

    def _section_categories(self, df_banque: pd.DataFrame) -> list:
        elements = [
            Paragraph("4. ANALYSE PAR CATÉGORIE", self.style_section),
            Spacer(1, 0.3*cm),
        ]

        analyse = (
            df_banque.groupby("categorie")["debit"]
            .sum()
            .sort_values(ascending=False)
            .head(10)
            .reset_index()
        )
        analyse.columns = ["Catégorie", f"Total Débits ({CONFIG['devise']})"]

        data = [list(analyse.columns)]
        for _, row in analyse.iterrows():
            data.append([str(row.iloc[0]), f"{float(row.iloc[1]):,.0f}"])

        table = Table(data, colWidths=[10*cm, 7*cm])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E86C1")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#EBF5FB"), colors.white]),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.5*cm))
        return elements

    def _pied_page(self) -> list:
        return [
            Spacer(1, 1*cm),
            HRFlowable(width="100%", thickness=1, color=colors.grey),
            Spacer(1, 0.2*cm),
            Paragraph(
                f"Document confidentiel — {CONFIG['entreprise']} | "
                f"SIRET : {CONFIG['siret']} | "
                f"Généré automatiquement le {datetime.now().strftime('%d/%m/%Y')}",
                ParagraphStyle("Pied", parent=self.style_normal,
                               fontSize=8, textColor=colors.grey,
                               alignment=TA_CENTER)
            ),
        ]


# ═══════════════════════════════════════════════════════════════════
# ORCHESTRATEUR PRINCIPAL
# ═══════════════════════════════════════════════════════════════════

class AutomatisationComptable:
    """
    Orchestrateur principal du pipeline d'automatisation comptable.
    Enchaîne parsing → import → catégorisation → rapprochement
    → détection anomalies → génération rapports.
    """

    def __init__(self):
        self.parser_pdf = ParserReleveBancairePDF()
        self.importeur = ImporteurDonneesComptables()
        self.api_bancaire = ImporteurAPIBancaire()
        self.categoriseur = CategoriseurTransactions()
        self.rapprocheur = RapprochementBancaire()
        self.detecteur = DetecteurAnomalies()
        self.generateur_excel = GenerateurRapportExcel()
        self.generateur_pdf = GenerateurRapportPDF()

        os.makedirs(CONFIG["dossier_entree"], exist_ok=True)
        os.makedirs(CONFIG["dossier_sortie"], exist_ok=True)

    def executer_pipeline(
        self,
        fichiers_pdf_banque: list = None,
        fichiers_comptables: list = None,
        utiliser_api: bool = False,
        compte_api: str = None,
        date_debut: datetime = None,
        date_fin: datetime = None,
    ) -> dict:
        """
        Exécute le pipeline complet de rapprochement.

        Args:
            fichiers_pdf_banque: Liste des chemins vers les relevés PDF
            fichiers_comptables: Liste des chemins vers les fichiers comptables
            utiliser_api: Utiliser l'API bancaire pour récupérer les transactions
            compte_api: Identifiant du compte bancaire (API)
            date_debut: Date de début de la période
            date_fin: Date de fin de la période

        Returns:
            dict avec les résultats et chemins des rapports
        """
        logger.info("=" * 65)
        logger.info("DÉMARRAGE DU PIPELINE D'AUTOMATISATION COMPTABLE")
        logger.info("=" * 65)

        date_debut = date_debut or (datetime.now().replace(day=1))
        date_fin = date_fin or datetime.now()
        periode = f"{date_debut.strftime('%d/%m/%Y')} — {date_fin.strftime('%d/%m/%Y')}"

        # ── Étape 1 : Import données bancaires ──────────────────────
        logger.info("\n[1/6] Import des données bancaires...")
        df_banque_list = []

        if fichiers_pdf_banque:
            for pdf in fichiers_pdf_banque:
                df = self.parser_pdf.parse_pdf(pdf)
                if not df.empty:
                    df_banque_list.append(df)

        if utiliser_api and compte_api:
            df_api = self.api_bancaire.recuperer_transactions(
                compte_api, date_debut, date_fin
            )
            if not df_api.empty:
                df_banque_list.append(df_api)

        df_banque = (
            pd.concat(df_banque_list, ignore_index=True)
            if df_banque_list else pd.DataFrame()
        )

        if df_banque.empty:
            logger.warning("Aucune donnée bancaire disponible. Création de données exemple.")
            df_banque = self._creer_donnees_exemple_banque()

        # ── Étape 2 : Import données comptables ──────────────────────
        logger.info("\n[2/6] Import des données comptables...")
        df_comptable_list = []

        if fichiers_comptables:
            for fichier in fichiers_comptables:
                df = self.importeur.importer_fichier(fichier)
                if df is not None and not df.empty:
                    df_comptable_list.append(df)

        df_comptable = (
            pd.concat(df_comptable_list, ignore_index=True)
            if df_comptable_list else pd.DataFrame()
        )

        if df_comptable.empty:
            logger.warning("Aucune donnée comptable. Création de données exemple.")
            df_comptable = self._creer_donnees_exemple_comptabilite()

        # ── Étape 3 : Catégorisation ──────────────────────────────────
        logger.info("\n[3/6] Catégorisation des transactions...")
        df_banque = self.categoriseur.categoriser_dataframe(df_banque)
        df_comptable = self.categoriseur.categoriser_dataframe(df_comptable)

        # ── Étape 4 : Rapprochement bancaire ─────────────────────────
        logger.info("\n[4/6] Rapprochement bancaire automatique...")
        resultats = self.rapprocheur.rapprocher(df_banque, df_comptable)

        # ── Étape 5 : Détection d'anomalies ──────────────────────────
        logger.info("\n[5/6] Détection des anomalies...")
        df_anomalies = self.detecteur.analyser(df_banque)

        # ── Étape 6 : Génération des rapports ─────────────────────────
        logger.info("\n[6/6] Génération des rapports...")
        horodatage = datetime.now().strftime("%Y%m%d_%H%M%S")

        chemin_excel = os.path.join(
            CONFIG["dossier_sortie"],
            f"rapprochement_bancaire_{horodatage}.xlsx"
        )
        chemin_pdf = os.path.join(
            CONFIG["dossier_sortie"],
            f"rapport_rapprochement_{horodatage}.pdf"
        )

        self.generateur_excel.generer(
            resultats, df_banque, df_comptable, df_anomalies, periode, chemin_excel
        )
        self.generateur_pdf.generer(
            resultats, df_banque, df_anomalies, periode, chemin_pdf
        )

        # ── Résumé final ───────────────────────────────────────────────
        logger.info("\n" + "=" * 65)
        logger.info("PIPELINE TERMINÉ AVEC SUCCÈS ✅")
        logger.info(f"  Taux rapprochement : {resultats['taux_rapprochement']:.1f}%")
        logger.info(f"  Anomalies détectées : {len(df_anomalies)}")
        logger.info(f"  Rapport Excel : {chemin_excel}")
        logger.info(f"  Rapport PDF   : {chemin_pdf}")
        logger.info("=" * 65)

        return {
            "df_banque": df_banque,
            "df_comptable": df_comptable,
            "resultats_rapprochement": resultats,
            "df_anomalies": df_anomalies,
            "chemin_excel": chemin_excel,
            "chemin_pdf": chemin_pdf,
            "taux_rapprochement": resultats["taux_rapprochement"],
            "periode": periode,
        }

    # ── Données de démonstration ────────────────────────────────────

    def _creer_donnees_exemple_banque(self) -> pd.DataFrame:
        """Crée des données bancaires de démonstration."""
        np.random.seed(42)
        n = 60
        dates = pd.date_range(
            start=datetime.now() - timedelta(days=30),
            periods=n, freq="12H"
        )

        libelles_debit = [
            "VIREMENT FOURNISSEUR CACAO GHANA LTD",
            "PAIEMENT SODECI FACTURE EAU USINE",
            "SALAIRES PERSONNEL MARS",
            "ACHAT EMBALLAGES PLASTIQUES SCI",
            "TRANSPORT LIVRAISON ABIDJAN-BOUAKE",
            "CARBURANT VEHICULE LIVRAISON TOTAL CI",
            "LOYER ENTREPOT ZONE INDUSTRIELLE PK24",
            "MAINTENANCE MACHINE ENSACHAGE",
            "PRIME ASSURANCE FLOTTE VEHICULE NSIA",
            "ACHAT ANACARDE COOPERATIVE KORHOGO",
            "FRAIS BANCAIRES SGBCI MARS",
            "TVA DGI VERSEMENT MENSUEL",
            "ABONNEMENT ORANGE INTERNET FIBRE",
            "ACHAT FARINE BLANCHE MOULIN MODERNE",
            "REPARATION GENERATEUR ELECTRICITE",
        ]

        libelles_credit = [
            "REGLEMENT CLIENT GRANDS MOULINS CI",
            "VIREMENT RECU EXPORT CACAO EUROPE",
            "PAIEMENT FACTURE SUPERMARCHE SOCOCE",
            "ACOMPTE COMMANDE CLIENT ABIDJAN FOOD",
            "REGLEMENT FACTURE GMS MARKET",
            "VIREMENT SUBVENTION FONDS AGRICOLE",
        ]

        transactions = []
        for i, date in enumerate(dates):
            est_credit = np.random.random() < 0.35
            if est_credit:
                libelle = np.random.choice(libelles_credit)
                montant = np.random.choice([
                    500_000, 1_200_000, 2_500_000,
                    750_000, 3_800_000, 1_500_000
                ]) + np.random.randint(-50_000, 50_000)
                debit, credit = 0, montant
            else:
                libelle = np.random.choice(libelles_debit)
                montant = np.random.choice([
                    200_000, 450_000, 1_800_000,
                    350_000, 800_000, 2_200_000
                ]) + np.random.randint(-30_000, 30_000)
                debit, credit = montant, 0

            transactions.append({
                "date": date,
                "libelle": libelle,
                "debit": debit,
                "credit": credit,
                "montant": credit - debit,
                "solde": None,
                "source": "Exemple",
                "id_transaction": f"BQ_{i+1:06d}",
            })

        df = pd.DataFrame(transactions)
        logger.info(f"Données exemple banque : {len(df)} transactions créées")
        return df

    def _creer_donnees_exemple_comptabilite(self) -> pd.DataFrame:
        """Crée des données comptables de démonstration."""
        np.random.seed(123)
        n = 55

        dates = pd.date_range(
            start=datetime.now() - timedelta(days=30),
            periods=n, freq="13H"
        )

        libelles = [
            "ACHAT MATIÈRES PREMIÈRES CACAO",
            "CHARGE SALARIALE PERSONNEL",
            "FACTURE SODECI EAU USINE",
            "VENTE PRODUITS FINIS CLIENT A",
            "TRANSPORT MARCHANDISES",
            "LOYER ENTREPÔT ZONE IND",
            "MAINTENANCE ÉQUIPEMENTS",
            "ASSURANCE FLOTTE NSIA",
            "ACHAT EMBALLAGES FOURNISSEUR",
            "RÈGLEMENT CLIENT EXPORT",
        ]

        ecritures = []
        for i, date in enumerate(dates):
            libelle = np.random.choice(libelles)
            est_charge = "ACHAT" in libelle or "CHARGE" in libelle or \
                         "TRANSPORT" in libelle or "LOYER" in libelle or \
                         "MAINTENANCE" in libelle or "ASSURANCE" in libelle
            montant = np.random.choice([
                150_000, 380_000, 750_000, 1_200_000, 2_000_000
            ]) + np.random.randint(-20_000, 20_000)
            debit = montant if est_charge else 0
            credit = 0 if est_charge else montant

            ecritures.append({
                "date": date + timedelta(hours=np.random.randint(0, 48)),
                "libelle": libelle,
                "debit": debit,
                "credit": credit,
                "montant": credit - debit,
                "compte": np.random.choice(["601000", "641000", "706000", "707000"]),
                "journal": np.random.choice(["AC", "VT", "BQ", "OD"]),
                "source_donnee": "Exemple",
                "id_ecriture": f"CPT_{i+1:06d}",
            })

        df = pd.DataFrame(ecritures)
        logger.info(f"Données exemple comptabilité : {len(df)} écritures créées")
        return df


# ═══════════════════════════════════════════════════════════════════
# POINT D'ENTRÉE
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 65)
    print("  AUTOMATISATION COMPTABLE — PME AGRO-ALIMENTAIRE")
    print("=" * 65)

    # ── Configuration de votre exécution ────────────────────────────
    # Modifiez ces paramètres selon votre situation :

    FICHIERS_PDF = [
        # "data/entree/releve_banque_mars_2025.pdf",
        # "data/entree/releve_banque_avril_2025.pdf",
    ]

    FICHIERS_COMPTABLES = [
        # "data/entree/export_sage_mars_2025.csv",
        # "data/entree/journal_achats.xlsx",
    ]

    DATE_DEBUT = datetime(2025, 3, 1)
    DATE_FIN = datetime(2025, 3, 31)

    # ── Lancement du pipeline ────────────────────────────────────────
    automatisation = AutomatisationComptable()

    resultats = automatisation.executer_pipeline(
        fichiers_pdf_banque=FICHIERS_PDF,
        fichiers_comptables=FICHIERS_COMPTABLES,
        utiliser_api=False,      # Passer à True si API configurée
        compte_api=None,
        date_debut=DATE_DEBUT,
        date_fin=DATE_FIN,
    )

    print(f"\n✅ PIPELINE TERMINÉ")
    print(f"   Taux de rapprochement : {resultats['taux_rapprochement']:.1f}%")
    print(f"   Anomalies détectées   : {len(resultats['df_anomalies'])}")
    print(f"\n📊 Rapport Excel : {resultats['chemin_excel']}")
    print(f"📄 Rapport PDF   : {resultats['chemin_pdf']}")